"""This module contains the main process of the robot."""

import os
import re
import json
from io import BytesIO
from typing import Literal, List
import time
import concurrent.futures

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from hvac import Client

from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from itk_dev_shared_components.graph import mail as graph_mail
from itk_dev_shared_components.graph import authentication as graph_authentication
from itk_dev_shared_components.smtp import smtp_util
from python_serviceplatformen import digital_post
from python_serviceplatformen.authentication import KombitAccess

from robot_framework import config


def process(orchestrator_connection: OrchestratorConnection) -> None:
    """ Do the primary process of the robot."""
    orchestrator_connection.log_trace("Running process.")
    process_arguments = json.loads(orchestrator_connection.process_arguments)

    # Access Keyvault
    vault_auth = orchestrator_connection.get_credential(config.KEYVAULT_CREDENTIALS)
    vault_uri = orchestrator_connection.get_constant(config.KEYVAULT_URI).value
    vault_client = Client(vault_uri)
    token = vault_client.auth.approle.login(role_id=vault_auth.username, secret_id=vault_auth.password)
    vault_client.token = token['auth']['client_token']

    # Get certificate
    read_response = vault_client.secrets.kv.v2.read_secret_version(mount_point='rpa', path=config.KEYVAULT_PATH, raise_on_deleted_version=True)
    certificate = read_response['data']['data']['cert']

    # Because KombitAccess requires a file, we save and delete the certificate after we use it
    certificate_path = "certificate.pem"
    with open(certificate_path, 'w', encoding='utf-8') as cert_file:
        cert_file.write(certificate)

    # Prepare access to service platform
    kombit_access = KombitAccess(process_arguments["service_cvr"], certificate_path, False)

    # Prepare access to email
    graph_credentials = orchestrator_connection.get_credential(config.GRAPH_API)
    graph_access = graph_authentication.authorize_by_username_password(graph_credentials.username, **json.loads(graph_credentials.password))
    mails = graph_mail.get_emails_from_folder(config.EMAIL_USER, config.MAIL_SOURCE_FOLDER, graph_access)

    for mail in mails:
        orchestrator_connection.log_trace("Reading emails.")
        # Get attachment from email
        attachments = graph_mail.list_email_attachments(mail, graph_access)
        email_attachment = graph_mail.get_attachment_data(attachments[0], graph_access)
        # Get data from email text
        requester = _get_recipient_from_email(mail.body)
        request_type = _get_request_type_from_email(mail.body)
        # Send and delete email
        start_time = time.time()
        return_data, rows_handled = handle_data(email_attachment, kombit_access, request_type, process_arguments["thread_count"])

        orchestrator_connection.log_info(f"{rows_handled} Rows handled. Total time spent: {time.time()-start_time} seconds")
        _send_status_email(requester, return_data)
        graph_mail.delete_email(mail, graph_access)


def handle_data(input_file: BytesIO, access: KombitAccess, service_type: Literal['Digital Post', 'NemSMS', 'Begge'], thread_count: int) -> tuple[BytesIO, int]:
    """ Read data from attachment, lookup each CPR number found and return a new file with added data.

    Args:
        input_file: Excel-file with rows of CPR to work on
        access: Kombit Access Token
        service_type: 'digitalpost', 'nemsms' or 'begge', to set which lookups to perform
    Returns:
        Filtered and formatted file with a list of people indicating whether they have Digital Post or not.
    """
    workbook = load_workbook(input_file)
    input_sheet: Worksheet = workbook.active

    # Check which services are requested
    service = ["Digital Post", "NemSMS"] if service_type == "Begge" else [service_type]

    # Call digital_post.is_registered for each input row and each required service
    data = threaded_service_check(input_sheet, service, access, thread_count)
    # data = linear_service_check(input_sheet, service, access)

    # Add data to excel sheet
    write_data_to_output_excel(service, data, input_sheet)

    # Grab workbook from memory and return it
    byte_stream = BytesIO()
    workbook.save(byte_stream)
    byte_stream.seek(0)
    return byte_stream, input_sheet.max_row


def threaded_service_check(input_sheet: Worksheet, service: List[str], kombit_access: KombitAccess, thread_count: int) -> dict[str, dict[str, bool]]:
    """ Call digital_post.is_registered for each input row and each required service.

    Args:
        input_sheet: The input worksheet containing rows of data.
        service: A list of services to check registration for.
        kombit_access: An object providing access credentials for the API.

    Returns:
        A dictionary with CPR as keys and lists of service registration results as values.
    """
    iter_ = iter(input_sheet)
    next(iter_)  # Skip header row

    data = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=thread_count) as executor:
        all_futures = {}
        for row in iter_:
            cpr = row[0].value  # Extract CPR from the row
            for s in service:
                service_type = s.replace(" ", "").lower()  # Format service name
                # Submit the API call to the thread pool
                future = executor.submit(digital_post.is_registered, cpr=cpr, service=service_type, kombit_access=kombit_access)
                all_futures[future] = {"cpr": cpr, "service_type": service_type}

        # Collect results as futures complete
        for future in concurrent.futures.as_completed(all_futures):
            cpr = all_futures[future]["cpr"]
            service_type = all_futures[future]["service_type"]
            if cpr not in data:
                data[cpr] = {}
            data[cpr][service_type] = future.result()  # Add the result to the corresponding CPR/service_type entry
    return data


def linear_service_check(input_sheet: Worksheet, service: List[str], kombit_access: KombitAccess) -> dict[str, dict[str, bool]]:
    """ Call digital_post.is_registered for each input row and each required service.

    Args:
        input_sheet: The input worksheet containing rows of data.
        service: A list of services to check registration for.
        kombit_access: An object providing access credentials for the API.

    Returns:
        A dictionary with CPR as keys and dictionary of service registration results as bools.
    """
    iter_ = iter(input_sheet)
    next(iter_)  # Skip header row

    data = {}
    for row in iter_:
        cpr = row[0].value  # Extract CPR from the row
        for s in service:
            serviceportal_type = s.replace(" ", "").lower()  # Format service name
            result = digital_post.is_registered(cpr=cpr, service=serviceportal_type, kombit_access=kombit_access)
            if cpr not in data:
                data[cpr] = {}
            data[cpr][serviceportal_type] = result

    return data


def write_data_to_output_excel(service: list[str], data: dict[str, dict[str, bool]], target_sheet: Worksheet) -> None:
    """ Add data to excel sheet.

    Args:
        service: Which services we add a status for
        data: A dictionary of id's with a list of booleans indicating if the id is registered with the service
        target_sheet: A sheet with id's in the first row
    """
    sheet_column_count = target_sheet.max_column
    # Write headers
    for col, s in enumerate(service, start=sheet_column_count + 1):  # Start from column 2 to avoid overwriting ID column
        # Write sheet headers for service types
        target_sheet.cell(row=1, column=col, value=s)
        # Go through rows of the sheet
        for row_idx, row in enumerate(target_sheet.iter_rows(min_row=2, max_col=1), start=2):
            cpr = row[0].value
            # Grab value of cpr-service_type from data dictionary and add to column
            status = data[cpr][s.replace(" ", "").lower()]
            status = "Tilmeldt" if status else "Ikke tilmeldt"
            target_sheet.cell(row=row_idx, column=col, value=status)


def _get_recipient_from_email(user_data: str) -> str:
    """ Find email in user_data using regex."""
    pattern = r"mailto:([^\"]+)"
    return re.findall(pattern, user_data)[0]


def _get_request_type_from_email(user_data: str) -> str:
    """ Find request type in user_data using regex."""
    pattern = r"Digital Post eller NemSMS<br>([^<]+)"
    return re.findall(pattern, user_data)[0]


def _send_status_email(recipient: str, file: BytesIO):
    """ Send an email to the requesting party and to the controller.

    Args:
        email: The email that has been processed.
    """
    smtp_util.send_email(
        recipient,
        config.EMAIL_STATUS_SENDER,
        "RPA: Udtræk om Tilmelding til Digital Post",
        "Robotten har nu udtrukket information om tilmelding til digital post i den forespurgte liste.\n\nVedhæftet denne mail finder du et excel-ark, som indeholder CPR-numre på navngivne borgere, for hvem robotten har slået op i Serviceplatformen og fået svar på, om de er tilmeldt digital post.\n\n Mvh. ITK RPA",
        config.SMTP_SERVER,
        config.SMTP_PORT,
        False,
        [smtp_util.EmailAttachment(file, config.EMAIL_ATTACHMENT)]
    )


if __name__ == '__main__':
    conn_string = os.getenv("OpenOrchestratorConnString")
    crypto_key = os.getenv("OpenOrchestratorKey")
    PROCESS_VARIABLES = r'{"service_cvr":"55133018", "thread_count":1}'
    oc = OrchestratorConnection("Udtræk Tilmelding Digital Post", conn_string, crypto_key, PROCESS_VARIABLES)
    process(oc)
